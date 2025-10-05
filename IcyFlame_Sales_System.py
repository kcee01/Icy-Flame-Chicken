import sqlite3
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook


class RestaurantSystem:
    def __init__(self, db_name="restaurant.db"):
        self.conn = sqlite3.connect(db_name)
        self.create_tables()

    def create_tables(self):
        with self.conn:
            self.conn.execute('''CREATE TABLE IF NOT EXISTS sales (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    item TEXT,
                                    price REAL,
                                    date TEXT,
                                    payment_type TEXT)''')

            self.conn.execute('''CREATE TABLE IF NOT EXISTS expenses (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    description TEXT,
                                    amount REAL,
                                    date TEXT)''')

    def record_sale(self, item, price, payment_type):
        with self.conn:
            self.conn.execute(
                "INSERT INTO sales (item, price, date, payment_type) VALUES (?, ?, ?, ?)",
                (item, price, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), payment_type.lower())
            )

    def record_expense(self, description, amount):
        with self.conn:
            self.conn.execute(
                "INSERT INTO expenses (description, amount, date) VALUES (?, ?, ?)",
                (description, amount, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            )

    def total_sales_by_payment(self):
        query = "SELECT payment_type, SUM(price) FROM sales GROUP BY payment_type"
        cursor = self.conn.execute(query)
        totals = {"cash": 0, "ewallet": 0, "orangemoney": 0, "smega": 0, "myzaka": 0}
        for payment_type, total in cursor:
            if payment_type in totals:
                totals[payment_type] = total
        return totals

    def total_expenses(self):
        cursor = self.conn.execute("SELECT SUM(amount) FROM expenses")
        total = cursor.fetchone()[0]
        return total if total else 0

    def cash_in_hand(self):
        totals = self.total_sales_by_payment()
        cash_sales = totals.get("cash", 0) or 0
        return cash_sales - self.total_expenses()

    def summary_text(self):
        lines = ["=== Daily Summary ===", "Sales by Payment Type:"]
        for method, total in self.total_sales_by_payment().items():
            lines.append(f"  {method}: {total}")
        lines.append(f"Total Expenses: {self.total_expenses()}")
        lines.append(f"Cash in Hand (cash sales - expenses): {self.cash_in_hand()}")
        return "\n".join(lines)

    def get_sales_today(self):
        today = datetime.now().strftime("%Y-%m-%d")
        cursor = self.conn.execute(
            "SELECT item, price, date, payment_type FROM sales WHERE date LIKE ?",
            (today + "%",)
        )
        return cursor.fetchall()

    def get_expenses_today(self):
        today = datetime.now().strftime("%Y-%m-%d")
        cursor = self.conn.execute(
            "SELECT description, amount, date FROM expenses WHERE date LIKE ?",
            (today + "%",)
        )
        return cursor.fetchall()

    def export_to_excel(self, filename="daily_summary.xlsx"):
        wb = Workbook()

        # ===== Summary Sheet =====
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Restaurant Daily Summary"])
        ws_summary.append([])
        ws_summary.append(["Sales by Payment Type"])
        for method, total in self.total_sales_by_payment().items():
            ws_summary.append([method, total])
        ws_summary.append([])
        ws_summary.append(["Total Expenses", self.total_expenses()])
        ws_summary.append(["Cash in Hand (cash-expenses)", self.cash_in_hand()])

        # ===== Sales Sheet =====
        ws_sales = wb.create_sheet(title="Sales Today")
        ws_sales.append(["Item", "Price", "Date", "Payment Type"])
        for row in self.get_sales_today():
            ws_sales.append(list(row))

        # ===== Expenses Sheet =====
        ws_expenses = wb.create_sheet(title="Expenses Today")
        ws_expenses.append(["Description", "Amount", "Date"])
        for row in self.get_expenses_today():
            ws_expenses.append(list(row))

        wb.save(filename)
        return filename


class RestaurantApp:
    def __init__(self, root):
        self.system = RestaurantSystem()
        self.root = root
        self.root.title("Restaurant Management System")

        tab_control = ttk.Notebook(root)

        # Sales Tab
        self.sales_tab = ttk.Frame(tab_control)
        tab_control.add(self.sales_tab, text="Record Sale")
        self.build_sales_tab()

        # Expenses Tab
        self.expenses_tab = ttk.Frame(tab_control)
        tab_control.add(self.expenses_tab, text="Record Expense")
        self.build_expenses_tab()

        # Summary Tab
        self.summary_tab = ttk.Frame(tab_control)
        tab_control.add(self.summary_tab, text="Summary")
        self.build_summary_tab()

        # Daily Records Tab
        self.records_tab = ttk.Frame(tab_control)
        tab_control.add(self.records_tab, text="Daily Records")
        self.build_records_tab()

        tab_control.pack(expand=1, fill="both")

    # ===== Build Sales Tab =====
    def build_sales_tab(self):
        ttk.Label(self.sales_tab, text="Item:").grid(row=0, column=0, padx=5, pady=5)
        self.item_entry = ttk.Entry(self.sales_tab)
        self.item_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self.sales_tab, text="Price:").grid(row=1, column=0, padx=5, pady=5)
        self.price_entry = ttk.Entry(self.sales_tab)
        self.price_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(self.sales_tab, text="Payment Type:").grid(row=2, column=0, padx=5, pady=5)
        self.payment_type = ttk.Combobox(
            self.sales_tab,
            values=["cash", "ewallet", "orangemoney", "smega", "myzaka"]
        )
        self.payment_type.grid(row=2, column=1, padx=5, pady=5)

        ttk.Button(self.sales_tab, text="Record Sale", command=self.add_sale).grid(row=3, column=0, columnspan=2, pady=10)

    # ===== Build Expenses Tab =====
    def build_expenses_tab(self):
        ttk.Label(self.expenses_tab, text="Description:").grid(row=0, column=0, padx=5, pady=5)
        self.expense_desc = ttk.Entry(self.expenses_tab)
        self.expense_desc.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self.expenses_tab, text="Amount:").grid(row=1, column=0, padx=5, pady=5)
        self.expense_amount = ttk.Entry(self.expenses_tab)
        self.expense_amount.grid(row=1, column=1, padx=5, pady=5)

        ttk.Button(self.expenses_tab, text="Record Expense", command=self.add_expense).grid(row=2, column=0, columnspan=2, pady=10)

    # ===== Build Summary Tab =====
    def build_summary_tab(self):
        self.summary_textbox = tk.Text(self.summary_tab, width=50, height=15)
        self.summary_textbox.pack(padx=10, pady=10)
        ttk.Button(self.summary_tab, text="Refresh Summary", command=self.show_summary).pack(pady=5)
        ttk.Button(self.summary_tab, text="Export to Excel", command=self.export_summary).pack(pady=5)

    # ===== Build Daily Records Tab =====
    def build_records_tab(self):
        ttk.Label(self.records_tab, text="Sales Today").pack(pady=5)
        self.sales_tree = ttk.Treeview(self.records_tab, columns=("Item", "Price", "Date", "Payment"), show="headings")
        for col in ("Item", "Price", "Date", "Payment"):
            self.sales_tree.heading(col, text=col)
            self.sales_tree.column(col, width=120)
        self.sales_tree.pack(padx=10, pady=5, fill="x")

        ttk.Label(self.records_tab, text="Expenses Today").pack(pady=5)
        self.expenses_tree = ttk.Treeview(self.records_tab, columns=("Description", "Amount", "Date"), show="headings")
        for col in ("Description", "Amount", "Date"):
            self.expenses_tree.heading(col, text=col)
            self.expenses_tree.column(col, width=120)
        self.expenses_tree.pack(padx=10, pady=5, fill="x")

        self.records_summary = tk.Text(self.records_tab, width=60, height=8)
        self.records_summary.pack(padx=10, pady=10)

        ttk.Button(self.records_tab, text="Refresh Records", command=self.show_records).pack(pady=5)

    # ===== Add Sale =====
    def add_sale(self):
        item = self.item_entry.get()
        try:
            price = float(self.price_entry.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid price")
            return
        payment = self.payment_type.get().lower()
        if item and payment:
            self.system.record_sale(item, price, payment)
            messagebox.showinfo("Success", "Sale recorded successfully")
            self.item_entry.delete(0, tk.END)
            self.price_entry.delete(0, tk.END)
            self.show_records()
        else:
            messagebox.showerror("Error", "Please fill all fields")

    # ===== Add Expense =====
    def add_expense(self):
        desc = self.expense_desc.get()
        try:
            amount = float(self.expense_amount.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid amount")
            return
        if desc:
            self.system.record_expense(desc, amount)
            messagebox.showinfo("Success", "Expense recorded successfully")
            self.expense_desc.delete(0, tk.END)
            self.expense_amount.delete(0, tk.END)
            self.show_records()
        else:
            messagebox.showerror("Error", "Please enter description")

    # ===== Show Summary =====
    def show_summary(self):
        self.summary_textbox.delete(1.0, tk.END)
        self.summary_textbox.insert(tk.END, self.system.summary_text())

    # ===== Export to Excel =====
    def export_summary(self):
        filename = self.system.export_to_excel()
        messagebox.showinfo("Exported", f"Summary exported to {filename}")

    # ===== Show Daily Records =====
    def show_records(self):
        for i in self.sales_tree.get_children():
            self.sales_tree.delete(i)
        for i in self.expenses_tree.get_children():
            self.expenses_tree.delete(i)

        for row in self.system.get_sales_today():
            self.sales_tree.insert("", tk.END, values=row)

        for row in self.system.get_expenses_today():
            self.expenses_tree.insert("", tk.END, values=row)

        self.records_summary.delete(1.0, tk.END)
        self.records_summary.insert(tk.END, self.system.summary_text())


if __name__ == "__main__":
    root = tk.Tk()
    app = RestaurantApp(root)
    root.mainloop()